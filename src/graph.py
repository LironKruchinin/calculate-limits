import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference

wb = Workbook(write_only=True)
ws = wb.create_sheet()
path = "output.xlsx"
wb_obj = openpyxl.load_workbook(path)
 
sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
 
# Loop will print all columns name
for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row = 1, column = i)
    print(cell_obj.value)
    
# rows = [
#     ('Number', 'Batch 1', 'Batch 2'),
#     (2, 10, 30),
#     (3, 40, 60),
#     (4, 50, 70),
#     (5, 20, 10),
#     (6, 10, 40),
#     (7, 50, 30),
# ]


# for row in rows:
#     ws.append(row)


# chart1 = BarChart()
# chart1.type = "col"
# chart1.style = 10
# chart1.title = "Bar Chart"
# chart1.y_axis.title = 'Test number'
# chart1.x_axis.title = 'Sample length (mm)'

# data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
# cats = Reference(ws, min_col=1, min_row=2, max_row=7)
# chart1.add_data(data, titles_from_data=True)
# chart1.set_categories(cats)
# chart1.shape = 4
# ws.add_chart(chart1, "A10")



wb.save("bar.xlsx")